#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出为Excel格式，提供更好的用户体验
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARNING] openpyxl未安装，将使用制表符分隔的Excel兼容格式")

def create_excel_format():
    """创建Excel格式的文件"""
    try:
        # 读取扫描结果
        with open('scan_results.txt', 'r', encoding='utf-8') as f:
            content = f.read()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f"磁盘分析报告_{timestamp}.xlsx"

        if OPENPYXL_AVAILABLE:
            return create_proper_excel(excel_file, content)
        else:
            return create_tab_delimited_excel(excel_file, content)

    except Exception as e:
        print(f"[ERROR] 创建Excel文件失败: {e}")
        return False, ""

def create_proper_excel(excel_file, content):
    """使用openpyxl创建真正的Excel文件"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "磁盘分析报告"

        # 定义样式
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        data_font = Font(size=11)

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 居中对齐
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 标题区域
        ws.merge_cells('A1:E1')
        ws['A1'] = "磁盘空间分析报告"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_alignment

        # 基本信息
        info_data = [
            ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["扫描路径", extract_scan_path(content)],
            ["扫描耗时", extract_scan_time(content)],
            ["总文件数", extract_total_files(content)],
            ["总大小", extract_total_size(content)]
        ]

        row = 3
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = data_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_alignment
            row += 1

        # 文件类型统计表
        row += 2
        ws[f'A{row}'] = "文件类型统计"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # 表头
        headers = ["文件类型", "文件数量", "占用大小", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1

        # 数据行
        file_types = extract_file_types(content)
        for file_type_data in file_types:
            for col, value in enumerate(file_type_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col == 1:  # 文件类型列
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            row += 1

        # 最大文件列表
        row += 2
        ws[f'A{row}'] = "最大文件列表"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = center_alignment
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # 表头
        file_headers = ["排名", "文件名", "大小", "路径"]
        for col, header in enumerate(file_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1

        # 数据行
        largest_files = extract_largest_files(content)
        for file_data in largest_files:
            for col, value in enumerate(file_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
            row += 1

        # 调整列宽
        column_widths = [15, 20, 15, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 保存文件
        wb.save(excel_file)
        return True, excel_file

    except Exception as e:
        print(f"[ERROR] 创建Excel文件失败: {e}")
        return False, ""

def create_tab_delimited_excel(excel_file, content):
    """创建制表符分隔的Excel兼容文件"""
    try:
        # 使用txt扩展名但格式为Excel兼容的制表符分隔
        txt_file = excel_file.replace('.xlsx', '_Excel兼容.txt')

        with open(txt_file, 'w', encoding='utf-8-sig') as f:
            f.write('\ufeff')  # BOM for Excel

            # 写入标题
            f.write("磁盘分析报告\n")
            f.write("=" * 50 + "\n\n")

            # 基本信息
            f.write("基本信息\n")
            f.write("-" * 20 + "\n")
            f.write(f"生成时间\t{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"扫描路径\t{extract_scan_path(content)}\n")
            f.write(f"扫描耗时\t{extract_scan_time(content)}\n")
            f.write(f"总文件数\t{extract_total_files(content)}\n")
            f.write(f"总大小\t{extract_total_size(content)}\n\n")

            # 文件类型统计
            f.write("文件类型统计\n")
            f.write("-" * 20 + "\n")
            f.write("文件类型\t文件数量\t占用大小\t占比\n")

            file_types = extract_file_types(content)
            for file_type_data in file_types:
                f.write("\t".join(file_type_data) + "\n")

            f.write("\n")

            # 最大文件列表
            f.write("最大文件列表\n")
            f.write("-" * 20 + "\n")
            f.write("排名\t文件名\t大小\t路径\n")

            largest_files = extract_largest_files(content)
            for file_data in largest_files:
                f.write("\t".join(file_data) + "\n")

        # 重命名为.csv以便Excel识别
        csv_file = txt_file.replace('_Excel兼容.txt', '.csv')
        os.rename(txt_file, csv_file)

        return True, csv_file

    except Exception as e:
        print(f"[ERROR] 创建Excel兼容文件失败: {e}")
        return False, ""

# 数据提取函数
def extract_scan_path(content):
    """提取扫描路径"""
    lines = content.split('\n')
    for line in lines:
        if "扫描路径:" in line:
            return line.split(":", 1)[1].strip()
    return "未知"

def extract_scan_time(content):
    """提取扫描时间"""
    lines = content.split('\n')
    for line in lines:
        if "扫描耗时:" in line:
            return line.split(":", 1)[1].strip()
    return "未知"

def extract_total_files(content):
    """提取总文件数"""
    lines = content.split('\n')
    for line in lines:
        if "符合条件的文件数:" in line or "扫描文件数:" in line:
            return line.split(":", 1)[1].strip()
    return "0"

def extract_total_size(content):
    """提取总大小"""
    lines = content.split('\n')
    for line in lines:
        if "总大小:" in line:
            return line.split(":", 1)[1].strip()
    return "0 B"

def extract_file_types(content):
    """提取文件类型统计"""
    file_types = []
    lines = content.split('\n')
    in_type_section = False

    for line in lines:
        if "文件类型统计:" in line:
            in_type_section = True
            continue
        elif "最大的文件" in line:
            in_type_section = False
            continue

        if in_type_section and line.strip() and ":" in line:
            try:
                parts = line.split(":")
                if len(parts) >= 2:
                    file_type = parts[0].strip()
                    details = parts[1].strip()
                    if "个文件" in details:
                        count_part = details.split("个文件")[0].strip()
                        size_part = details.split("个文件")[1].strip()
                        size = size_part.split("(")[0].strip()
                        percentage = size_part.split("(")[1].replace(")", "").strip() if "(" in size_part else ""

                        file_types.append([file_type, count_part, size, percentage])
            except:
                continue

    return file_types

def extract_largest_files(content):
    """提取最大文件列表"""
    largest_files = []
    lines = content.split('\n')
    in_file_section = False
    file_rank = 1

    for line in lines:
        if "最大的文件" in line:
            in_file_section = True
            continue
        elif "=" in line and in_file_section:
            in_file_section = False
            continue

        if in_file_section and line.strip() and (line.strip().endswith(".txt") or
            line.strip().endswith(".exe") or line.strip().endswith(".mp4") or
            line.strip().endswith(".zip") or line.strip().endswith(".pdf") or
            "." in line.strip()):
            try:
                if ". " in line and " - " in line:
                    parts = line.split(". ", 1)
                    if len(parts) == 2:
                        file_info = parts[1].split(" - ")
                        if len(file_info) == 2:
                            filename = file_info[0].strip()
                            filesize = file_info[1].strip()
                            largest_files.append([str(file_rank), filename, filesize, ""])
                            file_rank += 1
            except:
                continue

    return largest_files

def try_open_excel_with_file(file_path):
    """尝试用Excel打开文件"""
    try:
        os.startfile(file_path)
        return True
    except:
        return False

if __name__ == "__main__":
    success, file_path = create_excel_format()
    if success:
        print(f"[SUCCESS] Excel文件已创建: {file_path}")
        try_open_excel_with_file(file_path)
    else:
        print("[ERROR] Excel文件创建失败")